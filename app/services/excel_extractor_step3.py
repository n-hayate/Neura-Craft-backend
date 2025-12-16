from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List
import re

import openpyxl
from openpyxl.utils.datetime import from_excel

META_SHEET = "メタデータ"
LOG_SHEET = "LOG"
FORM_SHEET = "配合検討"


def _s(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _split_tags(s: str) -> list[str]:
    if not s:
        return []
    parts = re.split(r"[,\u3001/／|]+", s)
    return [p.strip() for p in parts if p and p.strip()]


def parse_step3_xlsx(xlsx_bytes: bytes) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(BytesIO(xlsx_bytes), data_only=True)

    meta = _parse_meta(wb)
    logs = _parse_log(wb)
    formulation = _parse_formulation(wb)

    derived = {
        "failure_tags": list(
            dict.fromkeys(
                _split_tags(_s(meta.get("failure_tags", "")))
                + [t for row in logs for t in _split_tags(_s(row.get("failure_symptoms", "")))]
            )
        ),
        "selected_variant": _s(meta.get("selected_variant", "")),
        "outcome": _s(meta.get("outcome", "")),
        "keywords": _split_tags(_s(meta.get("keywords", ""))),
    }

    return {
        "meta": meta,
        "log": logs,
        "formulation": formulation,
        "derived": derived,
        "template_version": "step3",
    }


def _parse_meta(wb) -> Dict[str, Any]:
    if META_SHEET not in wb.sheetnames:
        return {}
    ws = wb[META_SHEET]

    meta: Dict[str, Any] = {}
    # 3行目が ["項目","値","備考"]
    for r in range(4, 200):
        key = _s(ws.cell(r, 1).value)
        if not key:
            break
        val = ws.cell(r, 2).value
        if key == "date" and isinstance(val, (int, float)):
            try:
                meta[key] = str(from_excel(val).date())
            except Exception:
                meta[key] = _s(val)
        else:
            meta[key] = _s(val)
    return meta


def _parse_log(wb) -> List[Dict[str, Any]]:
    if LOG_SHEET not in wb.sheetnames:
        return []
    ws = wb[LOG_SHEET]

    headers: list[str] = []
    for c in range(1, 60):
        v = ws.cell(3, c).value
        if v is None:
            break
        headers.append(_s(v))

    rows: List[Dict[str, Any]] = []
    for r in range(4, 500):
        variant_id = _s(ws.cell(r, 1).value)
        if not variant_id:
            break

        d = {headers[i - 1]: _s(ws.cell(r, i).value) for i in range(1, len(headers) + 1)}

        # LLM入力安定用のキーを用意（日本語ヘッダ依存を減らす）
        d2 = {
            "variant_id": d.get("variant_id", variant_id),
            "variant_label": d.get("variant_label", ""),
            "purpose": d.get("目的（狙い）", ""),
            "change": d.get("変更点（前回比）", ""),
            "process_delta": d.get("工程条件の差分", ""),
            "eval_condition": d.get("評価条件（測定条件）", ""),
            "result": d.get("結果（数値＋所見）", ""),
            "judgement": d.get("判定（良/不良/要再検）", ""),
            "failure_symptoms": d.get("失敗症状（なければ「なし」）", ""),
            "cause_hypothesis": d.get("原因仮説（本文の範囲）", ""),
            "next_action": d.get("次アクション", ""),
            "quote": d.get("引用候補（短文）", ""),
            "keywords": d.get("関連キーワード", ""),
            "note": d.get("備考", ""),
        }
        rows.append(d2)
    return rows


def _parse_formulation(wb) -> Dict[str, Any]:
    if FORM_SHEET not in wb.sheetnames:
        return {"variants": [], "rows": []}
    ws = wb[FORM_SHEET]

    header_row = None
    for r in range(1, 80):
        if _s(ws.cell(r, 2).value) == "原材料 (銘柄)":
            header_row = r
            break
    if header_row is None:
        return {"variants": [], "rows": []}

    variants = []
    c = 3
    while c < 60:
        name = _s(ws.cell(header_row, c).value)  # "No. 1"
        if not name:
            break
        m = re.match(r"No\.\s*(\d+)", name)
        if m:
            variants.append({"variant_id": f"No.{m.group(1)}", "pct_col": c, "g_col": c + 1})
        c += 2

    start_row = header_row + 2  # データ開始
    rows = []
    for r in range(start_row, 500):
        ing = _s(ws.cell(r, 2).value)
        if not ing:
            break
        item = {"row_no": ws.cell(r, 1).value, "ingredient": ing, "variants": {}}
        for v in variants:
            item["variants"][v["variant_id"]] = {
                "pct": ws.cell(r, v["pct_col"]).value,
                "g": ws.cell(r, v["g_col"]).value,
            }
        rows.append(item)

    return {"variants": [v["variant_id"] for v in variants], "rows": rows}



