import asyncio
import logging
import os
import random
import sys
from datetime import datetime
from uuid import uuid4

from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from openpyxl import Workbook
from sqlalchemy.orm import Session

# Add project root to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.core.config import get_settings
from app.core.security import get_password_hash
from app.db.session import SessionLocal
from app.db.models.user import User
from app.db.models.file import File
from app.services.blob_service import BlobService

# Logging configuration
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Constants
DEMO_USER_EMAIL = "demo@example.com"
DEMO_USER_PASSWORD = "DemoUser123!"
ASSETS_DIR = "scripts/assets"

# Realistic Data Source
PRODUCTS = [
    "冷凍ハンバーグ", "鶏の唐揚げ", "ポテトサラダ", "豚骨ラーメンスープ", "カレールー", 
    "焼きおにぎり", "チーズケーキ", "ドレッシング（ごま）", "魚肉ソーセージ", "野菜ジュース"
]

ISSUES = [
    "肉汁不足", "焼成ロスが大きい", "退色", "食感が硬い", "分離する", 
    "味が薄い", "塩味が強い", "コストダウン", "歩留まり改善", "賞味期限延長"
]

INGREDIENTS = [
    "牛ミンチ", "玉ねぎ", "醤油", "増粘多糖類", "アミノ酸", 
    "鶏肉", "小麦粉", "乳化剤", "香料", "パプリカ色素"
]

CUSTOMERS = [
    "株式会社Aフーズ", "B商事", "スーパーC", "コンビニD", "E食品工業", 
    "Fダイニング", "G物産", "H製菓", "I乳業", "Jデリカ"
]

AUTHORS = ["開発 太郎", "研究 花子", "佐藤 健一", "鈴木 優子", "田中 実"]

def ensure_demo_user(db: Session) -> User:
    user = db.query(User).filter(User.email == DEMO_USER_EMAIL).first()
    if not user:
        logger.info(f"Creating demo user: {DEMO_USER_EMAIL}")
        user = User(
            email=DEMO_USER_EMAIL,
            full_name="Demo User",
            hashed_password=get_password_hash(DEMO_USER_PASSWORD),
            is_active=True
        )
        db.add(user)
        db.commit()
        db.refresh(user)
    else:
        logger.info(f"Using existing demo user: {DEMO_USER_EMAIL}")
    return user

def generate_pdf(filename: str, content: str) -> bytes:
    path = os.path.join(ASSETS_DIR, filename)
    os.makedirs(ASSETS_DIR, exist_ok=True)
    
    c = canvas.Canvas(path, pagesize=letter)
    c.drawString(100, 750, f"File: {filename}")
    c.drawString(100, 730, f"Content: {content}")
    c.drawString(100, 710, f"Generated at: {datetime.now()}")
    c.save()
    
    with open(path, "rb") as f:
        data = f.read()
    return data

def generate_excel(filename: str, content: str) -> bytes:
    path = os.path.join(ASSETS_DIR, filename)
    os.makedirs(ASSETS_DIR, exist_ok=True)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "File Name"
    ws["B1"] = filename
    ws["A2"] = "Content"
    ws["B2"] = content
    ws["A3"] = "Generated At"
    ws["B3"] = datetime.now()
    
    wb.save(path)
    
    with open(path, "rb") as f:
        data = f.read()
    return data

async def seed_data():
    db = SessionLocal()
    try:
        user = ensure_demo_user(db)
        blob_service = BlobService()
        
        # Clean up existing files for demo user? (Optional, keeping for now)
        
        logger.info("Starting data seeding...")
        
        for i in range(50):
            # Randomize Data
            product = random.choice(PRODUCTS)
            issue = f"{random.choice(ISSUES)} {random.choice(ISSUES)}" # 2 issues
            ingredient = f"{random.choice(INGREDIENTS)} {random.choice(INGREDIENTS)}"
            customer = random.choice(CUSTOMERS)
            author = random.choice(AUTHORS)
            trial_id = f"TR-2024-{str(i+1).zfill(3)}"
            
            ext = random.choice(["pdf", "xlsx"])
            original_filename = f"{product}_{trial_id}.{ext}"
            
            content_str = f"Product: {product}, Issue: {issue}, Customer: {customer}"
            
            if ext == "pdf":
                file_data = generate_pdf(original_filename, content_str)
                content_type = "application/pdf"
            else:
                file_data = generate_excel(original_filename, content_str)
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                
            # Upload to Blob
            blob_name, blob_url = blob_service.upload_bytes(
                file_data,
                original_filename,
                content_type
            )
            
            # Create DB Record
            file_obj = File(
                owner_id=user.id,
                original_filename=original_filename,
                content_type=content_type,
                file_size=len(file_data),
                blob_name=blob_name,
                azure_blob_url=blob_url,
                final_product=product,
                issue=issue,
                ingredient=ingredient,
                customer=customer,
                trial_id=trial_id,
                author=author,
                file_extension=ext,
                status="active"
            )
            db.add(file_obj)
            
            if (i + 1) % 10 == 0:
                logger.info(f"Created {i + 1} records...")
                
        db.commit()
        logger.info("Data seeding completed successfully!")
        
    except Exception as e:
        logger.error(f"Error seeding data: {e}")
        db.rollback()
    finally:
        db.close()

if __name__ == "__main__":
    asyncio.run(seed_data())

















