from uuid import uuid4
from datetime import datetime
import os
import random
import asyncio
import sys
from sqlalchemy.orm import Session

# Add project root to path
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.core.config import get_settings
from app.core.security import get_password_hash
from app.db.session import SessionLocal
from app.db.models.user import User
from app.services.blob_service import BlobService
from app.services.file_service import FileService
from app.schemas.file import FileCreate

# 作成したExcelジェネレーターをインポート
# NOTE: formula_generatorモジュールのパスに応じて調整してください
try:
    from formula_generator import generate_formula_excel
except ImportError:
    # フォールバック: モジュールが見つからない場合の処理
    def generate_formula_excel(product_name: str, trial_id: str, ingredients_pool: list) -> bytes:
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Product"
        ws["B1"] = product_name
        ws["A2"] = "Trial ID"
        ws["B2"] = trial_id
        ws["A3"] = "Ingredients"
        for i, ing in enumerate(ingredients_pool[:10], start=4):
            ws[f"A{i}"] = ing
        import io
        buffer = io.BytesIO()
        wb.save(buffer)
        return buffer.getvalue()

# メタデータのダミー
PRODUCTS = [
    "疑似果肉入りジャム", "低糖質クッキー", "高プロテインドリンク", "米粉パン", "ヴィーガンマヨネーズ",
    "カシューナッツパウダー", "アサイーボウル", "ピータン", "カロリーメイト"
    ]
ISSUES = [
    "粘度不足", "風味の劣化", "分離の発生", "コストダウン要求", "新規開発", 
    "触感改善", "加熱処理", "酸味抑制", "つや出し、色調"
    ]
INGREDIENTS = [
    "グラニュー糖", "パイナップル濃縮果汁", "ペクチンHM", "ペクチンLM", 
    "クエン酸", "クエン酸ナトリウム", "乳酸カルシウム", "難消化性デキストリン",
    "脱脂粉乳", "植物油脂", "加工澱粉", "香料", "アスコルビン酸","ペクチンUF、ファイバー", "ペクチン", "UNIPECTINE"
]
CUSTOMERS = ["ABC製菓", "XYZ飲料", "スーパーマーケットチェーンA", "コンビニエンスストアB","CCC食品", "山パン", "丸ケミ"]
AUTHORS = ["山田太郎", "鈴木一郎", "佐藤花子"]

# Constants
DEMO_USER_EMAIL = "user@example.com"
DEMO_USER_PASSWORD = "String123"


def ensure_demo_user(db: Session) -> User:
    """デモユーザーを取得または作成する"""
    user = db.query(User).filter(User.email == DEMO_USER_EMAIL).first()
    if not user:
        print(f"Creating demo user: {DEMO_USER_EMAIL}")
        user = User(
            email=DEMO_USER_EMAIL,
            full_name="Demo User",
            hashed_password=get_password_hash(DEMO_USER_PASSWORD),
            is_active=True
        )
        db.add(user)
        db.commit()
        db.refresh(user)
    else:
        print(f"Using existing demo user: {DEMO_USER_EMAIL}")
    return user


async def seed_data():
    db = SessionLocal()
    print("Seeding data started...")

    try:
        # デモユーザーを取得または作成
        user = ensure_demo_user(db)

        # BlobServiceとFileServiceの初期化
        async with BlobService() as blob_service:
            file_service = FileService(db)

            # 1. ベースとなるExcelファイルを1つ生成（これをコピーして使い回す）
            print("Generating base Excel template...")
            # ローカルにあるファイルを読み込む（存在する場合）
            base_file_path = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "scripts", "assets", "ダミーデータ_original.xlsx")
            if os.path.exists(base_file_path):
                print(f"Loading base template from: {base_file_path}")
                base_excel_data = open(base_file_path, 'rb').read()
            else:
                # ファイルが存在しない場合は生成関数を使用
                print("Base template file not found, generating new template...")
                base_excel_data = generate_formula_excel(
                    product_name="標準配合テンプレート", 
                    trial_id="TEMPLATE", 
                    ingredients_pool=INGREDIENTS
                )
            content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

            for i in range(10):
                # メタデータをランダムに選択
                product = random.choice(PRODUCTS)
                issue = random.choice(ISSUES)
                ingredient = random.choice(INGREDIENTS)
                customer = random.choice(CUSTOMERS)
                author = random.choice(AUTHORS)
                trial_id = f"TR-2024-{i+1:03d}"

                # ファイル名の生成ルール:
                # (application)_(issues)_(ingridient)_(customer)_(trial_id)_(auther).xlsx
                # ※ファイル名に使えない文字（スペース等）は適宜置換するなど運用に合わせて調整してください
                filename_base = f"{product}_{issue}_{ingredient}_{customer}_{trial_id}_{author}"
                original_name = f"{filename_base}.xlsx"

                # データは使い回し
                file_data = base_excel_data

                file_id = str(uuid4())
                # Blob保存名はUUIDのみにするか、元の拡張子を維持するか
                blob_name = f"{file_id}.xlsx"

                metadata = {
                    "original_name": original_name,
                    "status": "active",
                    "file_id": file_id,
                    "application": product,
                    "issue": issue,
                    "ingredient": ingredient,
                    "customer": customer,
                    "trial_id": trial_id,
                    "author": author,
                    "owner_id": str(user.id),  # Blobメタデータは文字列
                }

                print(f"Generating: {original_name}")

                # Blobアップロード
                blob_path, _ = await blob_service.upload_blob(
                    blob_name,
                    file_data,
                    content_type=content_type,
                    metadata=metadata,
                )

                # DBレコード作成
                payload = FileCreate(
                    id=file_id,
                    blob_path=blob_path,
                    original_name=original_name,
                    application=product,
                    issue=issue,
                    ingredient=ingredient,
                    customer=customer,
                    trial_id=trial_id,
                    author=author,
                    status="active",
                    is_preview_hidden=False,
                    owner_id=user.id,  # ← 整数型で設定
                )
                file_service.create(payload)

                if (i + 1) % 10 == 0:
                    print(f"Created {i + 1} records...")

        db.commit()
        print("Seeding completed successfully.")

    except Exception as e:
        print(f"Error seeding data: {e}")
        import traceback
        traceback.print_exc()
        db.rollback()
    finally:
        db.close()

if __name__ == "__main__":
    asyncio.run(seed_data())