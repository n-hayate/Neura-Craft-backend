"""
Create or update an Azure AI Search synonym map from `userdict.xlsx`.

Usage:
    export AZURE_SEARCH_ENDPOINT="https://<service>.search.windows.net"
    export AZURE_SEARCH_ADMIN_KEY="<admin-key>"  # or AZURE_SEARCH_API_KEY
    export AZURE_SEARCH_SYNONYM_MAP_NAME="userdict-synonyms-v1"  # optional
    python scripts/sync_synonym_map.py --xlsx data/userdict.xlsx

The script:
    1) reads the Excel rows as synonyms (one row -> one rule, comma-joined),
    2) normalizes terms (NFKC) by default,
    3) upserts the synonym map via SearchIndexClient.
"""

from __future__ import annotations

import argparse
import os
import sys
import unicodedata
from pathlib import Path
from typing import Iterable, List

from dotenv import load_dotenv

# Add project root to path for consistency with other scripts
sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from azure.core.credentials import AzureKeyCredential
from azure.search.documents.indexes import SearchIndexClient
from azure.search.documents.indexes.models import SynonymMap
from openpyxl import load_workbook


def env(name: str, default: str | None = None) -> str:
    value = os.getenv(name, default)
    if value is None:
        raise RuntimeError(f"Environment variable '{name}' is required.")
    return value


def normalize_term(term: str, do_normalize: bool = True) -> str:
    term = term.strip()
    if do_normalize:
        term = unicodedata.normalize("NFKC", term)
    return term


def iter_synonym_lines(
    *,
    xlsx_path: Path,
    skip_header: bool = True,
    normalize: bool = True,
) -> Iterable[str]:
    """Yield Solr-format synonym rules from the Excel file."""
    workbook = load_workbook(xlsx_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        for row_index, row in enumerate(sheet.iter_rows(values_only=True)):
            if skip_header and row_index == 0:
                continue

            terms: List[str] = []
            for cell in row:
                if cell is None:
                    continue
                text = normalize_term(str(cell), do_normalize=normalize)
                if text:
                    terms.append(text)

            # Remove duplicates while preserving order
            unique_terms = list(dict.fromkeys(terms))
            if len(unique_terms) >= 2:
                yield ", ".join(unique_terms)
    finally:
        workbook.close()


def upsert_synonym_map(
    *,
    endpoint: str,
    admin_or_query_key: str,
    map_name: str,
    synonyms: list[str],
) -> None:
    client = SearchIndexClient(endpoint=endpoint, credential=AzureKeyCredential(admin_or_query_key))
    synonym_map = SynonymMap(name=map_name, synonyms=synonyms)
    client.create_or_update_synonym_map(synonym_map)


def main() -> None:
    load_dotenv()
    parser = argparse.ArgumentParser(description="Create/Update Azure Search synonym map from Excel.")
    parser.add_argument("--xlsx", type=Path, default=Path("data/userdict.xlsx"), help="Path to synonym Excel file.")
    parser.add_argument("--map-name", type=str, default=os.getenv("AZURE_SEARCH_SYNONYM_MAP_NAME", "userdict-synonyms-v1"))
    parser.add_argument("--out", type=Path, default=Path("data/synonyms.solr.txt"), help="Path to write Solr rules.")
    parser.add_argument("--no-normalize", action="store_true", help="Disable NFKC normalization.")
    parser.add_argument("--no-skip-header", dest="skip_header", action="store_false", default=True, help="Do not skip the first row.")
    parser.add_argument("--export-only", action="store_true", help="Only export the .txt file without calling Azure.")
    args = parser.parse_args()

    if not args.xlsx.exists():
        raise FileNotFoundError(f"Excel file not found: {args.xlsx}")

    synonyms = list(
        iter_synonym_lines(
            xlsx_path=args.xlsx,
            skip_header=args.skip_header,
            normalize=not args.no_normalize,
        )
    )

    if not synonyms:
        raise RuntimeError("No synonym rules were generated. Check the Excel content and options.")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text("\n".join(synonyms), encoding="utf-8")
    print(f"Exported {len(synonyms)} synonym rules to {args.out}", flush=True)

    if args.export_only:
        print("Export-only mode. Skipped Azure Search update.", flush=True)
        return

    endpoint = env("AZURE_SEARCH_ENDPOINT")
    key = os.getenv("AZURE_SEARCH_ADMIN_KEY") or os.getenv("AZURE_SEARCH_API_KEY")
    if not key:
        raise RuntimeError("Provide AZURE_SEARCH_ADMIN_KEY or AZURE_SEARCH_API_KEY for synonym map update.")

    map_name = args.map_name or env("AZURE_SEARCH_SYNONYM_MAP_NAME")
    upsert_synonym_map(endpoint=endpoint, admin_or_query_key=key, map_name=map_name, synonyms=synonyms)
    print(f"Upserted synonym map '{map_name}' ({len(synonyms)} rules) to {endpoint}", flush=True)


if __name__ == "__main__":
    main()

